from time import sleep
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--headless=new")
driver = webdriver.Chrome(options=chrome_options)


dados = pd.DataFrame(columns=['Título', 'Preço', 'Link', 'Nota'])
print("Iniciando busca na Amazon...")
driver.get("https://www.amazon.com.br")
time.sleep(2)
try:
    busca = driver.find_element(By.CSS_SELECTOR, '[id="twotabsearchtextbox"]')
except:
    busca = driver.find_element(By.CSS_SELECTOR, '[id="nav-bb-search"]')
busca_realizada = input("Digite o termo de busca: ")
busca.send_keys(busca_realizada)
busca.send_keys(Keys.ENTER)
time.sleep(2)
try:
    proxima_pagina = driver.find_element(By.CSS_SELECTOR, 'a[aria-label*="Ir para a próxima página"]')
    while proxima_pagina != None:
        time.sleep(2)
        lista_elementos = driver.find_elements(By.CSS_SELECTOR, 'div[class="a-section a-spacing-base desktop-grid-content-view"]')
        for elemento in lista_elementos:
            titulo = elemento.find_element(By.CSS_SELECTOR, '[class="a-link-normal s-line-clamp-4 s-link-style a-text-normal"]').text
            try:
                preco = elemento.find_element(By.CSS_SELECTOR, 'span[class="a-price-whole"]').text
                centavos = elemento.find_element(By.CSS_SELECTOR, 'span[class="a-price-fraction"]').text
                preco = f'R$ {preco},{centavos}'
            except:
                preco = "Sem preço"
            link = elemento.find_element(By.CSS_SELECTOR, '[class="a-link-normal s-no-outline"]').get_attribute('href')
            nota = elemento.find_element(By.CSS_SELECTOR, 'div[data-cy="reviews-block"]>div>span').text
            print(f'Título: {titulo} | Preço: {preco} | Nota: {nota}')
            dados = pd.concat([dados, pd.DataFrame({'Título': [titulo], 'Preço': [preco], 'Link': [link], 'Nota': [nota]})], ignore_index=True)
        try:
            proxima_pagina = driver.find_element(By.CSS_SELECTOR, 'a[aria-label*="Ir para a próxima página"]')
            proxima_pagina.click()
        except:
            proxima_pagina = None
            break

        time.sleep(1)
except: 
    print("Não há mais páginas.")

data_atual = time.strftime("%d-%m-%Y_%H-%M-%S")
busca_realizada = f"{busca_realizada}_{data_atual}"
dados.to_excel(f'{busca_realizada}.xlsx', index=False)

# Adicionar hiperlinks

wb = load_workbook(f'{busca_realizada}.xlsx')
ws = wb.active

link_column = get_column_letter(3)

for row in range(2, len(dados) + 2):  # começa na linha 2 (pula cabeçalho)
    cell = ws[f'{link_column}{row}']
    link_value = cell.value
    if link_value:
        cell.hyperlink = link_value
        cell.value = "Clique aqui"  # ou deixe o link visível

wb.save(f'{busca_realizada}.xlsx')

driver.quit()
