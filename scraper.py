"""
Amazon Product Scraper
-----------------------
Scrapes product listings from Amazon and exports results to Excel.

Usage:
    python scraper.py --query "notebook gamer"
    python scraper.py --query "gaming laptop" --country us
    python scraper.py --query "clavier mecanique" --country fr --max-pages 3
    python scraper.py --query "auriculares" --country es --no-headless
    python scraper.py --query "monitor 4k" --country br --output resultados.xlsx
"""

import argparse
import logging
import random
import time
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

CONFIG_PATH = Path(__file__).parent / "config.yaml"


def load_config(path=CONFIG_PATH):
    """Load settings from config.yaml. Returns defaults if file is missing."""
    defaults = {
        "scraper": {
            "default_country": "br",
            "headless": True,
            "max_pages": None,
            "timeout": 10,
            "delay": {"min": 1.5, "max": 3.0},
        }
    }
    if not path.exists():
        logger.warning("config.yaml not found, using defaults.")
        return defaults
    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    merged = defaults.copy()
    merged["scraper"].update(data.get("scraper", {}))
    return merged


COUNTRIES = {
    "br": {"url": "https://www.amazon.com.br", "currency": "R$",  "decimal": ","},
    "us": {"url": "https://www.amazon.com",    "currency": "$",   "decimal": "."},
    "uk": {"url": "https://www.amazon.co.uk",  "currency": "GBP", "decimal": "."},
    "de": {"url": "https://www.amazon.de",     "currency": "EUR", "decimal": ","},
    "fr": {"url": "https://www.amazon.fr",     "currency": "EUR", "decimal": ","},
    "es": {"url": "https://www.amazon.es",     "currency": "EUR", "decimal": ","},
    "it": {"url": "https://www.amazon.it",     "currency": "EUR", "decimal": ","},
    "ca": {"url": "https://www.amazon.ca",     "currency": "CAD", "decimal": "."},
    "mx": {"url": "https://www.amazon.com.mx", "currency": "MXN", "decimal": "."},
    "jp": {"url": "https://www.amazon.co.jp",  "currency": "JPY", "decimal": "."},
}


class AmazonScraper:
    """Scrapes product listings from Amazon for any supported country."""

    USER_AGENTS = [
        (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0.0.0 Safari/537.36"
        ),
        (
            "Mozilla/5.0 (X11; Linux x86_64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
    ]

    def __init__(self, country="br", headless=False, max_pages=None,
                 timeout=10, delay_min=1.5, delay_max=3.0):
        if country not in COUNTRIES:
            raise ValueError(
                "Unsupported country '{}'. Choose from: {}".format(
                    country, ", ".join(COUNTRIES.keys())
                )
            )
        config = COUNTRIES[country]
        self.base_url = config["url"]
        self.currency = config["currency"]
        self.decimal_sep = config["decimal"]
        self.headless = headless
        self.max_pages = max_pages
        self.timeout = timeout
        self.delay_min = delay_min
        self.delay_max = delay_max
        self._query = ""
        self.driver = self._build_driver()
        self.wait = WebDriverWait(self.driver, self.timeout)
        logger.info("Country: %s | Site: %s | Currency: %s",
                    country.upper(), self.base_url, self.currency)

    def _build_driver(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        options.add_argument("user-agent={}".format(random.choice(self.USER_AGENTS)))
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        if self.headless:
            options.add_argument("--headless=new")
            options.add_argument("--window-size=1920,1080")
        driver = webdriver.Chrome(options=options)
        driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
        )
        return driver

    def search(self, query):
        """Navigate directly to the Amazon search results URL."""
        self._query = quote_plus(query)
        search_url = "{}/s?k={}".format(self.base_url, self._query)
        logger.info("Navigating to search URL: %s", search_url)
        self.driver.get(search_url)
        self._wait_for_results()

    def _wait_for_results(self):
        try:
            self.wait.until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, 'div[data-component-type="s-search-result"]')
                )
            )
        except TimeoutException:
            logger.warning("Results took too long to load; proceeding anyway.")

    def scrape_all_pages(self):
        """Collect products across all result pages using direct URL navigation."""
        all_products = []
        page = 1
        while True:
            logger.info("Scraping page %d...", page)
            products = self._scrape_current_page()
            all_products.extend(products)
            logger.info("  -> %d products found (total so far: %d)",
                        len(products), len(all_products))

            if self.max_pages and page >= self.max_pages:
                logger.info("Reached max-pages limit (%d).", self.max_pages)
                break

            if not self._has_next_page():
                logger.info("No more pages.")
                break

            page += 1
            self._human_delay()
            next_url = "{}/s?k={}&page={}".format(self.base_url, self._query, page)
            logger.info("Navigating to page %d...", page)
            self.driver.get(next_url)
            self._wait_for_results()

        return all_products

    def _has_next_page(self):
        """Check if a next page button exists (without clicking it)."""
        selectors = [
            "a.s-pagination-next",
            'a[aria-label*="next page"]',
            'a[aria-label*="proxima"]',
            'a[aria-label*="suivante"]',
            'a[aria-label*="nachste"]',
        ]
        for selector in selectors:
            try:
                self.driver.find_element(By.CSS_SELECTOR, selector)
                return True
            except NoSuchElementException:
                continue
        return False

    def _scrape_current_page(self):
        products = []
        cards = self.driver.find_elements(
            By.CSS_SELECTOR, 'div[data-component-type="s-search-result"]'
        )
        logger.info("  Cards found on page: %d", len(cards))
        for card in cards:
            product = self._extract_product(card)
            if product:
                products.append(product)
        return products

    def _extract_product(self, card):
        title = self._get_title(card)
        if not title:
            return None
        return {
            "Title":  title,
            "Price":  self._get_price(card, self.currency, self.decimal_sep),
            "Link":   self._get_link(card, self.base_url),
            "Rating": self._get_rating(card),
        }

    @staticmethod
    def _get_title(card):
        # Amazon wraps the <a> around the <h2>, so the title text lives in
        # `h2 span` (or the h2 itself). Exact full-class selectors break on any
        # markup change, so we rely on the stable h2 structure with fallbacks.
        selectors = ["h2 span", "h2", "[data-cy='title-recipe']"]
        for selector in selectors:
            try:
                element = card.find_element(By.CSS_SELECTOR, selector)
                text = (element.text or element.get_attribute("textContent") or "").strip()
                if text:
                    return text
            except NoSuchElementException:
                continue
        try:
            aria = card.find_element(By.CSS_SELECTOR, "h2").get_attribute("aria-label")
            if aria:
                return aria.strip()
        except NoSuchElementException:
            pass
        return ""

    @staticmethod
    def _get_link(card, base_url):
        # Most robust: every result card carries the product's ASIN.
        asin = card.get_attribute("data-asin")
        if asin:
            return "{}/dp/{}".format(base_url, asin)
        selectors = ["h2 a", "a.a-link-normal[href*='/dp/']", "a.a-link-normal.s-link-style"]
        for selector in selectors:
            try:
                href = card.find_element(By.CSS_SELECTOR, selector).get_attribute("href")
                if href:
                    return href
            except NoSuchElementException:
                continue
        return ""

    @staticmethod
    def _get_price(card, currency, decimal_sep):
        # Preferred: the fully formatted price Amazon hides in `.a-offscreen`
        # (e.g. "$599.00"). It is hidden text, so Selenium's `.text` returns "";
        # read `textContent` instead.
        try:
            offscreen = card.find_element(
                By.CSS_SELECTOR, "span.a-price span.a-offscreen"
            ).get_attribute("textContent")
            if offscreen and offscreen.strip():
                return offscreen.strip()
        except NoSuchElementException:
            pass
        # Fallback: reconstruct from the visible whole + fraction spans.
        try:
            whole = card.find_element(By.CSS_SELECTOR, "span.a-price-whole").text.strip()
            try:
                fraction = card.find_element(
                    By.CSS_SELECTOR, "span.a-price-fraction"
                ).text.strip()
            except NoSuchElementException:
                fraction = "00"
            return "{} {}{}{}".format(currency, whole, decimal_sep, fraction)
        except NoSuchElementException:
            return "No price"

    @staticmethod
    def _get_rating(card):
        # Rating sits in the star icon's alt text ("4.4 out of 5 stars"), which
        # is hidden (read via textContent), with localized aria-label fallbacks.
        selectors = [
            "i[class*='a-icon-star'] span.a-icon-alt",
            "span[aria-label*='out of 5 stars']",
            "span[aria-label*='de 5 estrelas']",
            "span[aria-label*='sur 5']",
            "span[aria-label*='von 5 Sternen']",
        ]
        for selector in selectors:
            try:
                element = card.find_element(By.CSS_SELECTOR, selector)
                text = (
                    element.get_attribute("textContent")
                    or element.get_attribute("aria-label")
                    or ""
                ).strip()
                if text:
                    return text.split(" ")[0]
            except NoSuchElementException:
                continue
        return "No rating"

    @staticmethod
    def export_to_excel(products, output_path):
        """Save products to an Excel file with clickable hyperlinks."""
        if not products:
            logger.warning("No products to save.")
            return
        df = pd.DataFrame(products)
        df.to_excel(output_path, index=False)
        wb = load_workbook(output_path)
        ws = wb.active
        link_col = get_column_letter(df.columns.get_loc("Link") + 1)
        for row in range(2, len(df) + 2):
            cell = ws["{}{}".format(link_col, row)]
            if cell.value:
                cell.hyperlink = cell.value
                cell.value = "Click here"
        wb.save(output_path)
        logger.info("Results saved to: %s", output_path)

    def _human_delay(self):
        time.sleep(random.uniform(self.delay_min, self.delay_max))

    def quit(self):
        self.driver.quit()


def parse_args():
    cfg = load_config()["scraper"]
    parser = argparse.ArgumentParser(
        description="Scrape product listings from Amazon and export to Excel.",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("--query", type=str, default=None,
                        help="Search term. Prompts if not provided.")
    parser.add_argument("--country", type=str, default=cfg["default_country"],
                        choices=list(COUNTRIES.keys()),
                        help="Amazon country to scrape (default from config.yaml).")
    parser.add_argument("--output", type=str, default=None,
                        help="Output .xlsx filename.")
    parser.add_argument("--max-pages", type=int, default=cfg["max_pages"],
                        metavar="N", help="Max pages to scrape.")
    parser.add_argument("--no-headless", action="store_true",
                        help="Show the Chrome browser window.")
    return parser.parse_args()


def main():
    cfg = load_config()["scraper"]
    args = parse_args()
    query = args.query or input("Enter search term: ").strip()
    if not query:
        raise SystemExit("Search term cannot be empty.")
    timestamp = time.strftime("%d-%m-%Y_%H-%M-%S")
    safe_query = query.replace(" ", "_")
    output = args.output or "{}_{}_{}.xlsx".format(safe_query, args.country, timestamp)
    headless = cfg["headless"] and not args.no_headless
    scraper = AmazonScraper(
        country=args.country,
        headless=headless,
        max_pages=args.max_pages,
        timeout=cfg["timeout"],
        delay_min=cfg["delay"]["min"],
        delay_max=cfg["delay"]["max"],
    )
    try:
        scraper.search(query)
        products = scraper.scrape_all_pages()
        scraper.export_to_excel(products, output)
        logger.info("Done! %d products exported.", len(products))
    finally:
        scraper.quit()


if __name__ == "__main__":
    main()
